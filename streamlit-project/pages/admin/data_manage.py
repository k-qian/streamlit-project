import streamlit as st
import openpyxl

from database import (
    get_childcare_centers,
    add_childcare_center,
    update_childcare_center,
    delete_childcare_center,
    bulk_insert_childcare_centers,
    get_exercise_videos,
    add_exercise_video,
    update_exercise_video,
    delete_exercise_video,
    bulk_insert_exercise_videos,
    get_music_recommendations,
    add_music_recommendation,
    update_music_recommendation,
    delete_music_recommendation,
    bulk_insert_music_recommendations,
)


def _parse_childcare_xlsx(file) -> list[tuple]:
    wb = openpyxl.load_workbook(file, read_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:  # name must exist
            rows.append((row[0] or "", row[1] or "", row[2] or "", row[3] or ""))
    wb.close()
    return rows


def _parse_exercise_xlsx(file) -> list[tuple]:
    wb = openpyxl.load_workbook(file, read_only=True)
    rows = []
    for name in wb.sheetnames:
        if name == "運動建議":
            ws = wb[name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    rows.append((row[0], row[1]))
    wb.close()
    return rows


def _parse_music_xlsx(file) -> list[tuple]:
    wb = openpyxl.load_workbook(file, read_only=True)
    rows = []
    for name in wb.sheetnames:
        if name == "音樂建議":
            ws = wb[name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    rows.append((row[0], row[1]))
    wb.close()
    return rows


def _childcare_tab():
    st.subheader("托嬰中心管理")

    # --- Upload ---
    with st.expander("📤 匯入 xlsx"):
        uploaded = st.file_uploader(
            "上傳托嬰中心 xlsx", type=["xlsx"], key="cc_upload",
        )
        mode = st.radio(
            "匯入模式", ["清空後匯入", "追加匯入"], horizontal=True, key="cc_mode",
        )
        if uploaded and st.button("確認匯入", key="cc_import"):
            rows = _parse_childcare_xlsx(uploaded)
            if rows:
                count = bulk_insert_childcare_centers(
                    rows, clear_first=(mode == "清空後匯入"),
                )
                st.success(f"成功匯入 {count} 筆托嬰中心資料")
                st.rerun()
            else:
                st.warning("檔案中沒有有效資料")

    # --- Add ---
    with st.expander("➕ 新增單筆"):
        with st.form("cc_add"):
            col1, col2 = st.columns(2)
            with col1:
                new_type = st.text_input("機構類型", key="cc_new_type")
                new_name = st.text_input("機構名稱", key="cc_new_name")
            with col2:
                new_addr = st.text_input("地址", key="cc_new_addr")
                new_phone = st.text_input("電話", key="cc_new_phone")
            if st.form_submit_button("新增", use_container_width=True):
                if new_name:
                    add_childcare_center(new_type, new_name, new_addr, new_phone)
                    st.success("新增成功")
                    st.rerun()
                else:
                    st.warning("機構名稱為必填")

    # --- List ---
    centers = get_childcare_centers()
    st.caption(f"共 {len(centers)} 筆資料")

    if not centers:
        st.info("尚無資料，請匯入或新增")
        return

    for c in centers:
        with st.container(border=True):
            col_info, col_actions = st.columns([4, 1])
            with col_info:
                st.markdown(f"**{c['name']}**　`{c['type']}`")
                st.caption(f"📍 {c['address']}　📞 {c['phone']}")
            with col_actions:
                col_edit, col_del = st.columns(2)
                with col_edit:
                    if st.button("✏️", key=f"cc_edit_{c['id']}"):
                        st.session_state[f"cc_editing_{c['id']}"] = True
                with col_del:
                    if st.button("🗑️", key=f"cc_del_{c['id']}"):
                        delete_childcare_center(c["id"])
                        st.rerun()

            # Inline edit form
            if st.session_state.get(f"cc_editing_{c['id']}"):
                with st.form(f"cc_edit_form_{c['id']}"):
                    e_col1, e_col2 = st.columns(2)
                    with e_col1:
                        e_type = st.text_input("機構類型", value=c["type"])
                        e_name = st.text_input("機構名稱", value=c["name"])
                    with e_col2:
                        e_addr = st.text_input("地址", value=c["address"])
                        e_phone = st.text_input("電話", value=c["phone"])
                    col_save, col_cancel = st.columns(2)
                    with col_save:
                        if st.form_submit_button("儲存", use_container_width=True):
                            update_childcare_center(c["id"], e_type, e_name, e_addr, e_phone)
                            del st.session_state[f"cc_editing_{c['id']}"]
                            st.rerun()
                    with col_cancel:
                        if st.form_submit_button("取消", use_container_width=True):
                            del st.session_state[f"cc_editing_{c['id']}"]
                            st.rerun()


def _exercise_tab():
    st.subheader("運動影片管理")

    # --- Upload ---
    with st.expander("📤 匯入 xlsx"):
        uploaded = st.file_uploader(
            "上傳孕婦各種建議 xlsx（讀取「運動建議」工作表）",
            type=["xlsx"], key="ev_upload",
        )
        mode = st.radio(
            "匯入模式", ["清空後匯入", "追加匯入"], horizontal=True, key="ev_mode",
        )
        if uploaded and st.button("確認匯入", key="ev_import"):
            rows = _parse_exercise_xlsx(uploaded)
            if rows:
                count = bulk_insert_exercise_videos(
                    rows, clear_first=(mode == "清空後匯入"),
                )
                st.success(f"成功匯入 {count} 筆運動影片")
                st.rerun()
            else:
                st.warning("檔案中沒有有效的運動建議資料")

    # --- Add ---
    with st.expander("➕ 新增單筆"):
        with st.form("ev_add"):
            new_title = st.text_input("影片名稱", key="ev_new_title")
            new_url = st.text_input("YouTube 網址", key="ev_new_url")
            if st.form_submit_button("新增", use_container_width=True):
                if new_title and new_url:
                    add_exercise_video(new_title, new_url)
                    st.success("新增成功")
                    st.rerun()
                else:
                    st.warning("名稱和網址為必填")

    # --- List ---
    videos = get_exercise_videos()
    st.caption(f"共 {len(videos)} 筆資料")

    if not videos:
        st.info("尚無資料，請匯入或新增")
        return

    for v in videos:
        with st.container(border=True):
            col_info, col_actions = st.columns([4, 1])
            with col_info:
                st.markdown(f"**{v['title']}**")
                st.caption(v["url"])
            with col_actions:
                col_edit, col_del = st.columns(2)
                with col_edit:
                    if st.button("✏️", key=f"ev_edit_{v['id']}"):
                        st.session_state[f"ev_editing_{v['id']}"] = True
                with col_del:
                    if st.button("🗑️", key=f"ev_del_{v['id']}"):
                        delete_exercise_video(v["id"])
                        st.rerun()

            if st.session_state.get(f"ev_editing_{v['id']}"):
                with st.form(f"ev_edit_form_{v['id']}"):
                    e_title = st.text_input("影片名稱", value=v["title"])
                    e_url = st.text_input("YouTube 網址", value=v["url"])
                    col_save, col_cancel = st.columns(2)
                    with col_save:
                        if st.form_submit_button("儲存", use_container_width=True):
                            update_exercise_video(v["id"], e_title, e_url)
                            del st.session_state[f"ev_editing_{v['id']}"]
                            st.rerun()
                    with col_cancel:
                        if st.form_submit_button("取消", use_container_width=True):
                            del st.session_state[f"ev_editing_{v['id']}"]
                            st.rerun()


def _music_tab():
    st.subheader("音樂建議管理")

    # --- Upload ---
    with st.expander("📤 匯入 xlsx"):
        uploaded = st.file_uploader(
            "上傳孕婦各種建議 xlsx（讀取「音樂建議」工作表）",
            type=["xlsx"], key="mr_upload",
        )
        mode = st.radio(
            "匯入模式", ["清空後匯入", "追加匯入"], horizontal=True, key="mr_mode",
        )
        if uploaded and st.button("確認匯入", key="mr_import"):
            rows = _parse_music_xlsx(uploaded)
            if rows:
                count = bulk_insert_music_recommendations(
                    rows, clear_first=(mode == "清空後匯入"),
                )
                st.success(f"成功匯入 {count} 筆音樂建議")
                st.rerun()
            else:
                st.warning("檔案中沒有有效的音樂建議資料")

    # --- Add ---
    with st.expander("➕ 新增單筆"):
        with st.form("mr_add"):
            new_title = st.text_input("音樂名稱", key="mr_new_title")
            new_url = st.text_input("YouTube 網址", key="mr_new_url")
            if st.form_submit_button("新增", use_container_width=True):
                if new_title and new_url:
                    add_music_recommendation(new_title, new_url)
                    st.success("新增成功")
                    st.rerun()
                else:
                    st.warning("名稱和網址為必填")

    # --- List ---
    items = get_music_recommendations()
    st.caption(f"共 {len(items)} 筆資料")

    if not items:
        st.info("尚無資料，請匯入或新增")
        return

    for m in items:
        with st.container(border=True):
            col_info, col_actions = st.columns([4, 1])
            with col_info:
                st.markdown(f"**{m['title']}**")
                st.caption(m["url"])
            with col_actions:
                col_edit, col_del = st.columns(2)
                with col_edit:
                    if st.button("✏️", key=f"mr_edit_{m['id']}"):
                        st.session_state[f"mr_editing_{m['id']}"] = True
                with col_del:
                    if st.button("🗑️", key=f"mr_del_{m['id']}"):
                        delete_music_recommendation(m["id"])
                        st.rerun()

            if st.session_state.get(f"mr_editing_{m['id']}"):
                with st.form(f"mr_edit_form_{m['id']}"):
                    e_title = st.text_input("音樂名稱", value=m["title"])
                    e_url = st.text_input("YouTube 網址", value=m["url"])
                    col_save, col_cancel = st.columns(2)
                    with col_save:
                        if st.form_submit_button("儲存", use_container_width=True):
                            update_music_recommendation(m["id"], e_title, e_url)
                            del st.session_state[f"mr_editing_{m['id']}"]
                            st.rerun()
                    with col_cancel:
                        if st.form_submit_button("取消", use_container_width=True):
                            del st.session_state[f"mr_editing_{m['id']}"]
                            st.rerun()


def data_manage_page():
    if st.session_state.username != "admin":
        st.error("⛔ 無權限存取此頁面")
        return

    st.caption("管理後台")
    st.title("資料管理")

    tab_cc, tab_ev, tab_mr = st.tabs(["🏠 托嬰中心", "🏃 運動影片", "🎵 音樂建議"])

    with tab_cc:
        _childcare_tab()

    with tab_ev:
        _exercise_tab()

    with tab_mr:
        _music_tab()
